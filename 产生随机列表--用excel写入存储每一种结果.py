import time
import random
import gc
#_______excel操作___________可分写入和提取___
from openpyxl import Workbook
import datetime
#wb2= load_workbook("C:\\Users\\Administrator\\Desktop\\新建 Microsoft Excel 工作表.xlsx")
wb = Workbook() #创建一个workbook
ws = wb.active  #对wb的具体操作
#ws1 = wb.create_sheet("Myshee1")
#ws2 = wb2.active
#_____________继承excel__________________
#从excel里面提取值,返回一个二级列表
def ev(a,b):
    for i in range(a):
        for j in range(b):
            pass
    pass

#____________求阶乘__________________________
def jiec(self):
    if self == 0 or self == 1:
        jienum = 1
    else:
        jienum = self * jiec(self - 1)
    return jienum
#_____________________________产生全排列，传入一个1~n列表的长度__________________________________________
def sjlb(n=5,list3=[],list3_1=[],list4=[],count1=0):
    print('执行函数：sjlb')
    nj =jiec(n)
    for i3 in range(n):
        list3_1.append(i3)

    while count1!=nj:
        #random.shuffle所返回的值他的统计学规律是什么？每一种可能都是一样的分布吗？
        random.shuffle(list3_1)
        list3=list3_1[:]
        #not in 查找的时间复杂度有点大
        if (list3 not in list4):
            list4.append(list3)
            ws.append(list3)
            count1 += 1
            print('第{0}种可能{1}'.format(count1, list3))
            print('进度{0}/{1},{2:.2f}%'.format(count1, nj, ((count1 / nj) * 100)))

        # del list3
        # gc.collect()

    print('sjlb执行完毕！')
    print('全排列共有n！次中可能，本次共有{0}个元素，所以有{1}种结果'.format(n,count1))
    #存储excel表格
    wb.save('9种全排列穷举.xlsx')
    return list4
if __name__ == '__main__':
    startime=time.perf_counter()
    n=9
    nj=jiec(n)

    listl = sjlb(n)

    print(listl)

    listl.sort()

    print('排序后：',listl)

    endtime=time.perf_counter()
    dur=endtime-startime
    print('程序运行时间：',dur)
    print('time.process_time()程序运行时间：',time.process_time())


